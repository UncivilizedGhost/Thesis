import pandas as pd
import os
import glob
import hashlib
from openpyxl import load_workbook
from datetime import datetime

 

# ── Auth ──────────────────────────────────────────────────────────────────────

def get_user() -> str:
    """
    Prompt for username. If admin, verify password
    """
    attempts = 0
    while attempts < 3:
        username = input("Enter username: ").strip()
        if username == "admin":
            password = input("Enter password: ").strip()
            hashed_input = hashlib.sha256(password.encode()).hexdigest()

            if hashed_input == os.environ.get("PASSWORD", ""):
                print("[DEBUG] get_user: admin authenticated successfully")
                return "admin"
            else:
                print("[DEBUG] get_user: incorrect password")
                attempts += 1
        else:
            return username
    print("Too many failed attempts, falling back to regular user")
    return input("Enter username: ").strip()


# ── Admin tools ───────────────────────────────────────────────────────────────

async def list_excel_files() -> dict:
    """
    List all Excel (.xlsx) files in the current folder and their worksheet names.
    Returns
    -------
    dict
        Maps filename -> list of worksheet names.
    """
    result = {}
    for filepath in glob.glob("*.xlsx"):
        filename = os.path.basename(filepath)
        if filename.startswith("~$"): #skip temp files
            continue
        try:
            wb = load_workbook(filepath, read_only=True)
            result[filename] = wb.sheetnames
            wb.close()
        except Exception as e:
            result[filename] = [f"Error: {e}"]
    return result


async def get_timetable(file_name: str, sheet_name: str) -> str:
    """
    Get the schedule from an Excel worksheet.
    Parameters
    ----------
    file_name : str
    sheet_name : str
    Returns
    -------
    str
        String representation of the timetable.
    """
    df = pd.read_excel(file_name, sheet_name=sheet_name)
    result = (df.to_string())
    return result


async def clear_worksheet(file_name: str, sheet_name: str) -> str:
    """
    Clear all bookings in a worksheet, keeping row 1 column A.

    Parameters
    ----------
    file_name : str
    sheet_name : str

    Returns
    -------
    str
        Confirmation message.
    """
    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            if cell.value is not None:
                cell.value = None

    wb.save(file_name)
    return f"Cleared {sheet_name} in {file_name}"


async def read_log_files() -> str:
    """
    Read all .log files in the current folder and return their combined contents.

    Returns
    -------
    str
        All log contents labelled by filename.
    """

    logs = []
    files = sorted(glob.glob("*.log"))
    if not files:
        return "No log files found."
    for filepath in files:
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                content = f.read().strip()

            logs.append(f"=== {filepath} ===\n{content}")
        except Exception as e:
            logs.append(f"=== {filepath} ===\nError: {e}")
    return "\n\n".join(logs)


# ── Equipment ─────────────────────────────────────────────────────────────────

async def get_additive_manufacturing_equipment() -> dict:
    """
    Retrieve additive manufacturing equipment with booking requirements and hourly cost.


    Returns
    -------
    dict
        Categories mapped to lists of {name, description, requires_booking, cost}.
        cost is € per hour. requires_booking is a bool.
    """

    db = {
        'printers': [
            {'name': 'FDM 3D Printer',         'description': 'Extrudes thermoplastics layer by layer.',       'requires_booking': True,  'cost': 10},
            {'name': 'SLA 3D Printer',          'description': 'UV laser cures liquid resin into solid parts.', 'requires_booking': True,  'cost': 15},
            {'name': 'Metal 3D Printer (DMLS)', 'description': 'Laser sinters metal powder for metal parts.',   'requires_booking': True,  'cost': 50},
        ],
        'design': [
            {'name': 'SolidWorks',   'description': 'Parametric 3D CAD software.',       'requires_booking': False, 'cost': 0},
            {'name': 'Fusion 360',   'description': 'Cloud-based CAD/CAM platform.',      'requires_booking': False, 'cost': 0},
            {'name': 'AutoCAD',      'description': '2D and 3D drafting software.',       'requires_booking': False, 'cost': 0},
            {'name': 'Cura',         'description': 'Slicing software, exports G-code.',  'requires_booking': False, 'cost': 0},
            {'name': 'PrusaSlicer',  'description': 'Advanced slicing software.',         'requires_booking': False, 'cost': 0},
            {'name': 'PreForm',      'description': 'Formlabs SLA preparation software.', 'requires_booking': False, 'cost': 0},
        ],
        'post_processing': [
            {'name': 'Support Removal Tools',    'description': 'Pliers and snips for supports.', 'requires_booking': False, 'cost': 0},
            {'name': 'Sanding & Polishing Tools','description': 'Surface finishing tools.',        'requires_booking': False, 'cost': 0},
            {'name': 'UV Curing Station',        'description': 'Post-cures resin prints.',        'requires_booking': True,  'cost': 5},
            {'name': 'Paints and Coating',       'description': 'Airbrush and spray finishing.',   'requires_booking': False, 'cost': 0},
        ],
        'quality': [
            {'name': 'Calipers',                 'description': 'Dimensional measurement.',      'requires_booking': False, 'cost': 0},
            {'name': 'Micrometers',              'description': 'High-precision measurement.',    'requires_booking': False, 'cost': 0},
            {'name': '3D Scanner',               'description': 'Geometry verification.',         'requires_booking': True,  'cost': 5},
            {'name': 'Surface Roughness Tester', 'description': 'Surface quality measurement.',   'requires_booking': False, 'cost': 0},
        ],
    }
    return db

# ── Booking ───────────────────────────────────────────────────────────────────


DAY_COL = {
    "monday": "B", "tuesday": "C", "wednesday": "D",
    "thursday": "E", "friday": "F", "saturday": "G", "sunday": "H",
}

 

async def get_available_slots(
    file_name: str,
    sheet_name: str,
    duration_hours: int,
) -> str:
    """
    Return all available consecutive blocks of the required duration (06:00-22:00).
    The agent reads this list and picks the best slot based on user preferences.
    Obtain file_name and sheet_name first by calling list_excel_files.

    Parameters
    ----------
    file_name : str
    sheet_name : str
    duration_hours : int

    Returns
    -------
    str
        All free slots, one per line: "<day> <HH:00>-<HH:00>"
        Or "NO_SLOTS_AVAILABLE" if none found.
    """
    print(f"[DEBUG] get_available_slots: file='{file_name}' sheet='{sheet_name}' duration={duration_hours}h")


    try:
        wb = load_workbook(file_name, read_only=True)
    except Exception as e:
        return f"Error loading file: {e}"

    if sheet_name not in wb.sheetnames:
        wb.close()
        return f"Worksheet named '{sheet_name}' not found"

    ws = wb[sheet_name]
    available = []

    for day_key, col in DAY_COL.items():
        for start_hour in range(6, 22):
            if start_hour + duration_hours > 22:
                break
            
            all_free = True
            
            # Check every hour block 
            for hour in range(start_hour, start_hour + duration_hours):
                row = hour - 4
                cell_val = ws[f"{col}{row}"].value
                
                # If cell is not empy, it's booked
                if cell_val is not None and str(cell_val).strip() != "":
                    all_free = False
                    break
            
            if all_free:
                end_hour = start_hour + duration_hours
                slot_str = f"{day_key} {start_hour:02d}:00-{end_hour:02d}:00"
                available.append(slot_str)

    wb.close()

    if not available:
        return "NO_SLOTS_AVAILABLE"

    return "\n".join(available)



async def add_booking(
    day: str,
    start_time: str,
    end_time: str,
    name: str,
    file_name: str,
    sheet_name: str,
) -> str:
    """
    Book a multi-hour slot in the schedule.
    Writes name into every hour from start_time up to (not including) end_time.
    Obtain file_name and sheet_name first by calling list_excel_files.

    Parameters
    ----------
    day : str
        Lowercase day name (e.g. 'monday').
    start_time : str
        Start hour in HH:00 format (e.g. "18:00").
    end_time : str
        End hour in HH:00 format (e.g. "21:00"). Exclusive.
    name : str
        Username to write into each slot.
    file_name : str
        Excel filename. Obtain from list_excel_files.
    sheet_name : str
        Worksheet name. Obtain from list_excel_files.

    Returns
    -------
    str
        "Success" or an error message.
    """
    if day not in DAY_COL:
        msg = f"Invalid day: '{day}'. Choose from: {', '.join(DAY_COL)}"

        return msg

    start_hour = int(start_time.split(":")[0])
    end_hour   = int(end_time.split(":")[0])

    if start_hour < 6 or end_hour > 22 or start_hour >= end_hour:
        msg = f"Invalid range {start_time}-{end_time}. Allowed: 06:00-22:00."
        return msg

    col = DAY_COL[day]
    wb  = load_workbook(file_name)
    ws  = wb[sheet_name]

    # Check all cells are free before writing any (atomic)
    for hour in range(start_hour, end_hour):
        row  = hour - 4
        cell = ws[f"{col}{row}"]
        if cell.value not in (None, ""):
            msg = f"Slot {day} {hour:02d}:00 already booked by '{cell.value}'. No changes made."
            return msg

    # Write username into every hour in the range
    for hour in range(start_hour, end_hour):
        row        = hour - 4
        cell       = ws[f"{col}{row}"]
        cell.value = name

    wb.save(file_name)
    return "Success"


# ── Session logging ───────────────────────────────────────────────────────────

from datetime import datetime

def write_session_log(
    username: str, 
    user_request: str, 
    approved_plan: str, 
    bookings: list[dict], 
    total_cost: int, 
    skipped: list[str]
) -> str:
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"session_{username}_{timestamp}.log"

    lines = [
        "SESSION LOG",
        "===========",
        f"Timestamp : {timestamp}",
        f"User      : {username}",
        f"Request   : {user_request}",
        "",
        "APPROVED PLAN",
        "-------------", 
        approved_plan,
        "",
        "BOOKINGS",
        "--------"
    ]

    if bookings:
            for b in bookings:
                tool   = b.get('tool', 'Unknown Tool')
                day    = b.get('day', '?')
                start  = b.get('start_time', '?')
                end    = b.get('end_time', '?')
                status = b.get('status', '?')
                
                log_entry = f"  {tool}: {day} {start}-{end} [{status}]"
                lines.append(log_entry)
    else:
        lines.append("  None")

    if skipped:
        lines += ["", "SKIPPED", "-------"]
        lines += [f"  {s}" for s in skipped]

    # 4. Finish the log
    lines += [
        "",
        f"TOTAL COST : €{total_cost}"
    ]

    # Write it all to the file
    with open(filename, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    return filename