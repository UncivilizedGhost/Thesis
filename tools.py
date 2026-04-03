import pandas as pd



async def get_timetable()-> str:
    """
    Get current schedule 

    Returns
    -------
    str
        A string containing the timetable
    """

    df = pd.read_excel("test.xlsx", sheet_name="Sheet1")
    return str(df)











async def get_additive_manufacturing_equipment(equipment_type: str = None) -> dict:
    """
    Retrieve additive manufacturing equipment categorized by function, including
    booking requirements and cost metadata.

    This function provides a structured inventory of tools and systems used in
    additive manufacturing workflows, spanning equipment for fabrication,
    design, post-processing, and quality assurance.

    Parameters
    ----------
    equipment_type : str, optional
        Specifies which category of equipment to retrieve. Valid options are:

        - 'printers' :
            Additive manufacturing machines (e.g., FDM, SLA, metal printers).
                    
        - 'design' :
            CAD and slicing software used for modeling and print preparation.
        
        - 'post_processing' :
            Tools used after printing for finishing and refining parts.
        
        - 'quality' :
            Measurement and inspection tools for validating printed parts.
        
        - None (default) :
            Returns all categories.

    Returns
    -------
    dict
        A dictionary mapping equipment categories to lists of equipment items.

        Each equipment item is represented as a dictionary with the following fields:

        - 'name' : str
            The name of the equipment or software.

        - 'description' : str
            A concise explanation of the tool's purpose and functionality.

        - 'requires_booking' : bool
            Indicates whether the equipment must be reserved before use.
            Typically True for shared or limited-access resources (e.g., printers).

        - 'cost' : int
            Indicates the cost of using the device

        Example structure:
        {
            'printers': [
                {
                    'name': 'Fused Deposition Modeling (FDM) 3D Printers',
                    'description': '...',
                    'requires_booking': True,
                    'cost': 1
                },
                ...
            ],
            'design': [...],
            'post_processing': [...],
            'quality': [...]
        }

    Raises
    ------
    ValueError
        If `equipment_type` is provided but is not one of the valid categories:
    """    
    equipment_database = {
        'printers': [
            {
                'name': 'Fused Deposition Modeling (FDM) 3D Printers',
                'description': 'Extrude thermoplastics like PLA, ABS, and PETG layer by layer.',
                'requires_booking': True,
                'cost': 1
            },
            {
                'name': 'Stereolithography (SLA) 3D Printers',
                'description': 'Use UV laser to cure liquid resin into solid parts.',
                'requires_booking': True,
                'cost': 1
            },
            {
                'name': 'Metal 3D Printers (DMLS)',
                'description': 'Direct Metal Laser Sintering for metal components.',
                'requires_booking': True,
                'cost': 1
            }
        ],
        
        'design': [
            {
                'name': 'CAD Software - SolidWorks',
                'description': 'Professional parametric 3D modeling software.',
                'requires_booking': False,
                'cost': 0
            },
            {
                'name': 'CAD Software - Fusion 360',
                'description': 'Cloud-based CAD/CAM platform.',
                'requires_booking': False,
                'cost': 0
            },
            {
                'name': 'CAD Software - AutoCAD',
                'description': '2D and 3D drafting software.',
                'requires_booking': False,
                'cost': 0
            },
            {
                'name': 'Slicing Software - Cura',
                'description': 'Converts 3D models into G-code.',
                'requires_booking': False,
                'cost': 0
            },
            {
                'name': 'Slicing Software - PrusaSlicer',
                'description': 'Advanced slicing software.',
                'requires_booking': False,
                'cost': 0
            },
            {
                'name': 'Slicing Software - PreForm',
                'description': 'Formlabs SLA preparation software.',
                'requires_booking': False,
                'cost': 0
            }
        ],
        
        'post_processing': [
            {
                'name': 'Support Removal Tools',
                'description': 'Pliers, cutters, and snips.',
                'requires_booking': False,
                'cost': 0
            },
            {
                'name': 'Sanding & Polishing Tools',
                'description': 'Surface finishing tools.',
                'requires_booking': False,
                'cost': 0
            },
            {
                'name': 'UV Curing Stations',
                'description': 'Post-cure resin prints.',
                'requires_booking': False,
                'cost': 0
            },
            {
                'name': 'Paints and Coating Equipment',
                'description': 'Final finishing tools.',
                'requires_booking': False,
                'cost': 0
            }
        ],
        
        'quality': [
            {
                'name': 'Calipers',
                'description': 'Dimensional measurement tools.',
                'requires_booking': False,
                'cost': 0
            },
            {
                'name': 'Micrometers',
                'description': 'High-precision measurement tools.',
                'requires_booking': False,
                'cost': 0
            },
            {
                'name': '3D Scanners',
                'description': 'Geometry verification tools.',
                'requires_booking': False,
                'cost': 0
            },
            {
                'name': 'Surface Roughness Testers',
                'description': 'Surface quality measurement.',
                'requires_booking': False,
                'cost': 0
            }
        ]
    }

    valid_types = ['printers', 'design', 'post_processing', 'quality']

    if equipment_type is None:
        return equipment_database
    elif equipment_type in valid_types:
        return {equipment_type: equipment_database[equipment_type]}
    else:
        raise ValueError(f"Invalid equipment_type. Must be one of: {', '.join(valid_types)} or None")










from openpyxl import load_workbook
from openpyxl.styles import PatternFill

async def add_booking(day:str, start_time:str, end_time:str, name:str )-> str:
    """
    Book a time slot on a weekly schedule for a specific person.
    
    This function writes a person's name into a timesheet Excel file for a
    specified day and time range, effectively reserving that time slot.
    
    Parameters
    ----------
    day : str
        The day of the week for the booking. Valid options are:
        - 'monday', 'tuesday', 'wednesday', 'thursday', 'friday', 
          'saturday', 'sunday'
    
    start_time : str
        The start time of the booking in 24-hour format (e.g., "09:00", "14:30")
    
    end_time : str
        The end time of the booking in 24-hour format (e.g., "10:00", "15:30")
    
    name : str
        The name of the person booking the time slot
    Notes
    -----
    - Time slots are divided into 30-minute increments
    - The function will overwrite any existing names in the specified time range
    - The Excel file is saved after each booking operation
    """
    file_path="test.xlsx"
    sheet_name="Sheet1"

    
    wb = load_workbook(file_path)
    ws = wb[sheet_name]



    week_map = {
   
    "monday": "B",
    "tuesday": "C",
    "wednesday": "D",
    "thursday": "E",
    "friday": "F",
    "saturday": "G",
    "sunday": "H",
    }


    column=week_map[day]

    def time_string_to_row(t):
        hour, minute = map(int, t.split(":"))
        return (hour * 2 + minute // 30) + 2

    start_row=time_string_to_row(start_time)
    end_row=time_string_to_row(end_time)

    for row in range(start_row, end_row + 1):
        cell = ws[f"{column}{row}"]
        cell.value = name

    wb.save(file_path)

    return "Success"


#fill_and_style("test.xlsx", "Sheet1", "monday", "10:00", "11:00", "Alice")
